import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from db import connect_to_mongo


def create_excel_file(collection_name):
    db = connect_to_mongo()
    if db is None:
        print("Failed to connect to MongoDB.")
        return

    collection = db[f'{collection_name}']
    documents = collection.find()

    data = []
    for doc in documents:
        website = doc.get('query', '')
        name = website.split("//")[-1].split("/")[0].replace('www.', '') if website else ''
        for suffix in ['.co.uk', '.com', '.net', '.org', '.biz', '.eco']:
            name = name.replace(suffix, '')
        name = name.replace('-', ' ').title() if name else ''
        email = doc.get('emails', [{'value': ''}])[0]['value'] if 'emails' in doc else ''
        phone = doc.get('phones', [{'value': ''}])[0]['value'] if 'phones' in doc else ''
        country = doc.get('details', {}).get('country', '') if 'details' in doc else ''
        city = doc.get('details', {}).get('city', '') if 'details' in doc else ''
        socials = doc.get('socials', {}) if 'socials' in doc else {}
        instagram = socials.get('instagram', '')
        facebook = socials.get('facebook', '')
        twitter = socials.get('twitter', '')
        youtube = socials.get('youtube', '')
        linkedin = socials.get('linkedin', '')
        about = doc.get('site_data', {}).get('description', '') if 'site_data' in doc else ''
        classification = doc.get('classification', {})
        category = classification.get('category', '')
        # subcategory = classification.get('subcategory', '')
        weight = doc.get('weight', '')

        data.append({
            'NAME': name,
            'SERVICE OFFERED': category,
            'WEBSITE': website,
            'EMAIL': email,
            'PHONE': phone,
            'COUNTRY': country,
            'CITY': city,
            'INSTAGRAM ACCOUNT (LINK)': instagram,
            'FACEBOOK ACCOUNT (LINK)': facebook,
            'TWITTER ACCOUNT (LINK)': twitter,
            'YOUTUBE ACCOUNT (LINK)': youtube,
            'LINKEDIN ACCOUNT (LINK)': linkedin,
            'ABOUT': about,
            'MAX WEIGHT (%)': weight
        })

    if not data:
        print("No data found in the collection.")
        return

    df = pd.DataFrame(data)
    print("DataFrame created with the following data:")
    print(df)

    output_file_path = f'/Users/hakanakyuz/Downloads/{collection_name}.xlsx'
    df.to_excel(output_file_path, index=False)

    wb = load_workbook(output_file_path)
    ws = wb.active

    column_widths = {
        'A': 25,
        'B': 25,
        'C': 40,
        'D': 40,
        'E': 20,
        'F': 15,
        'G': 20,
        'H': 45,
        'I': 45,
        'J': 45,
        'K': 45,
        'L': 45,
        'M': 25,
        'N': 8,
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        ws.row_dimensions[row[0].row].height = 30
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)

    wb.save(output_file_path)

    print(f"Data saved successfully to {output_file_path}.")
